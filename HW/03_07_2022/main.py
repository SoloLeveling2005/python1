import openpyxl
import psycopg2 as psycopg2
from openpyxl import Workbook
from openpyxl.utils import get_column_interval
from openpyxl.utils import get_column_letter

file_name = "main.xlsx"

row = 2
col = 1

global_mass = []

workbook = openpyxl.load_workbook(file_name)
worksheet = workbook.active
while True:
    mass_local = []
    while True:

        value = worksheet.cell(row=row, column=col).value
        if value is None:
            col = 1
            break
        print(value)
        col += 1
        mass_local.append(value)

    global_mass.append(mass_local)
    row += 1
    value = worksheet.cell(row=row, column=col).value
    if value is None:
        break

host = "127.0.0.1"
user = "postgres"
password = "Solo2005"
db_name = "new_postgres_db"

try:
    # connect to exist database

    connection = psycopg2.connect(
        host=host,
        user=user,
        password=password,
        database=db_name
    )
    connection.autocommit = True
    for data in global_mass:
        print(data)
        with connection.cursor() as cursor:
            cursor.execute(
                f"""INSERT INTO public.new_postgres_db ("age ", is_have_money, name) VALUES ('{data[0]}', '{data[1]}', '{data[2]}');""")




    print("[INFO] Data was succefully inserted")

except Exception as _ex:
    print("[INFO] Error while working with PostgreSQL", _ex)
finally:
    if connection:
        # cursor.close()
        connection.close()
        print("[INFO] PostgreSQL connection closed")

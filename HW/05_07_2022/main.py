# -*- coding: utf-8 -*-

import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_interval
from openpyxl.utils import get_column_letter

host = "127.0.0.1"
user = "postgres"
password = "Solo2005"
db_name = "new1_postgres_db"
import psycopg2
import time

import threading

import requests
import json

url = "https://jsonplaceholder.typicode.com/posts"

response = requests.get(url)
json_data = response.content.decode()

i = 0
while True:
    try:
        j = json.loads(json_data)
        j_userId = j[i]["userId"]
        j_id = j[i]["id"]
        j_name = j[i]["title"]
        j_body = j[i]["body"]

        connection = psycopg2.connect(
            host=host,
            user=user,
            password=password,
            database=db_name
        )
        connection.autocommit = True
        i += 1
        with connection.cursor() as cursor:
            print(j_id)
            cursor.execute(
                f"""INSERT INTO new1_postgres_tb (
                        "userId", id, title, body)
                        VALUES ('{j_userId}', '{j_id}', '{str(j_name)}', '{j_body}');"""
            )

            print("[INFO] Data was succefully inserted")
    except:
        print("Ой. Список закончился.")
        break

mass = []


def select_email_by_id():
    try:

        # connect to exist database

        connection = psycopg2.connect(
            host=host,
            user=user,
            password=password,
            database=db_name
        )
        connection.autocommit = True
        userId = 5
        try:
            local_mass = []

            with connection.cursor() as cursor:
                cursor.execute(
                    f"""SELECT title
                            FROM public.new1_postgres_tb WHERE "userId" = {userId} ;"""
                )

                result_title = cursor.fetchall()
            local_mass.append(result_title)
            with connection.cursor() as cursor:
                cursor.execute(
                    f"""SELECT body
                            FROM public.new1_postgres_tb WHERE "userId" = {userId} ;"""
                )

                result_body = cursor.fetchall()


            local_mass.append(result_body)
            # print("res", result_title)
            # print("res", result_body)
            # print("local_mass", local_mass)
            global mass
            mass.append(local_mass)

        finally:
            print("Ой. Список закончился.")


    except Exception as _ex:
        print("[INFO] Error while working with PostgresSQL", _ex)
    finally:
        if connection:
            # cursor.close()
            connection.close()
            print("[INFO] PostgresSQL connection closed")


select_email_by_id()

print("mass",mass)
print("mass0",mass[0][0])
print("mass1",mass[0][1])

workbook2 = Workbook()
worksheet2 = workbook2.active
worksheet2.column_dimensions['A'].width = 80
worksheet2.column_dimensions['B'].width = 300


def insert_result():
    i = 1
    for data in mass[0][0]:
        print(mass[0][0])
        print(data)
        # row = mass.index(data) + 1

        worksheet2.cell(row=i, column=1).value = str(data[0])

        i += 1
    i = 1
    for data in mass[0][1]:
        # row = mass.index(data) + 1
        print(data)
        worksheet2.cell(row=i, column=2).value = str(data[0])

        i += 1


insert_result()

workbook2.save("data.xlsx")

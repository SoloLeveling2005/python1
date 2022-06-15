from random import *
# import openpyxl
import openpyxl
from openpyxl import Workbook

# создадим рандомные данные в exel файл


workbook = Workbook()
worksheet = workbook.active

worksheet.cell(row=1, column=1).value = "Имя"
worksheet.cell(row=1, column=2).value = "Фамилия"
worksheet.cell(row=1, column=3).value = "Должность"

list_first_name = ["Александр", "Михаил", "Максим", "Даниил", "Лев", "Артем", "Марк", "Иван", "Дмитрий", "Матвей",
                   "Андрей", "Никита", "Егор", "Алексей", "Арсений", "Константин", "Давид", "Сергей"]
list_second_name = ["Смирнов", "Иванов", "Кузнецов", "Соколов", "Попов", "Лебедев", "Козлов", "Новиков", "Морозов",
                    "Петров", "Волков", "Соловьёв"]
list_work = ["junior", "middle", "senior"]

for row in range(1, 100):
    worksheet.cell(row=row, column=1).value = choice(list_first_name)
    worksheet.cell(row=row, column=2).value = choice(list_second_name)
    worksheet.cell(row=row, column=3).value = choice(list_work)

workbook.save("first_doc.xlsx")

# создадим рандомные данные в exel файл


# переносим данные

name_file = "first_doc.xlsx"
workbook = openpyxl.load_workbook(name_file)
worksheet = workbook.active


class Main:
    def __init__(self):
        self.global_list = []
        row = 1
        col = 1
        while True:
            local_list = []
            while True:
                value = worksheet.cell(row=row, column=col).value
                if value is None:
                    break
                local_list.append(value)
                col += 1
            self.global_list.append(local_list)
            row += 1
            col = 1
            value = worksheet.cell(row=row, column=col).value
            if value is None:
                break


results = Main()
results = results.global_list

workbook2 = Workbook()
worksheet2 = workbook2.active

index = 0


class Create:
    def __init__(self, index):
        for row in results:
            print(row)
            for col in row:
                print(col)
                index_col = row.index(col)
                # index_row = results.index(row)
                worksheet2.cell(row=index + 1, column=index_col + 1).value = str(col)
            index += 1


go = Create(index)

workbook2.save("result.xlsx")

# переносим данные

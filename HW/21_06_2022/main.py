# -*- coding: utf-8 -*-

import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_interval
from openpyxl.utils import get_column_letter


glob_list = [[],[],[]]

row_list = [2,4,6]
file_name = "data.xlsx"

print(glob_list)


workbook = openpyxl.load_workbook(file_name)
worksheet = workbook.active
for row in row_list:
    col = 1
    while True:
        value = worksheet.cell(row=row, column=col).value
        if value is None:
            break
        print(row)
        glob_list[row_list.index(row)].append(value)
        # value = value.split("_")
        # if value[0] == "a":
        #     glob_list[0].append(value)
        # elif value[0] == "b":
        #     glob_list[1].append(value)
        # elif value[0] == "c":
        #     glob_list[2].append(value)
        # print(value)

        col += 1

print(glob_list)

a_elements = set(glob_list[0])
b_elements = set(glob_list[1])
c_elements = set(glob_list[2])
# print(a_elements)
# print(b_elements)
# z = a_elements.intersection(b_elements)
# print(z)
# print()
print("Пересечение a и b элементов: "+str(a_elements.intersection(b_elements)))
print("Пересечение a и с элементов: "+str(a_elements.intersection(c_elements)))

print("Разница a и b элементов: "+str(a_elements.difference(b_elements)))
print("Разница a и с элементов: "+str(a_elements.difference(c_elements)))


print("Дополняем b элементами c: "+str(b_elements.union(c_elements)))
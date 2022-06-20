import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_interval
from openpyxl.utils import get_column_letter

first_file = "Лист1.xlsx"
second_file = "Лист2.xlsx"
third_file = "Лист3.xlsx"
first_file_list_global = []
second_file_list_global = []
third_file_list_global = []
files = [[first_file, first_file_list_global], [second_file, second_file_list_global],
         [third_file, third_file_list_global]]



# создаем глобальный список в который загрузим все данные со всех файлов
glob_list = []

for i in files:
    workbook = openpyxl.load_workbook(i[0])
    worksheet = workbook.active
    for y in range(1, 26):
        mass_local = []
        for x in range(1, 26):
            value = worksheet.cell(row=x, column=y).value
            if value is None:
                break
            mass_local.append(value)
        glob_list.append(mass_local)

    def iii():
        for ii in glob_list:
            if len(ii) == 0:
                del glob_list[0]
                iii()


# записываем в новый файл
workbook2 = Workbook()
worksheet2 = workbook2.active


print(glob_list)
row = 0
col = 1
coll = 1
for el in glob_list:
    print("el" + str(el))
    col = get_column_letter(coll)
    print(col)
    if len(el) == 0:
        coll -= 1
    for le in el:
        row += 1
        worksheet2[f'{col}{row}'] = str(le)
        print(le)
    coll += 1
    row = 0

workbook2.save("task1.xlsx")
# записываем в новый файл
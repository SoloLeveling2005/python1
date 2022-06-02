import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_interval
from openpyxl.utils import get_column_letter

first_file = "Лист1.xlsx"
second_file = "Лист2.xlsx"
third_file = "Лист3.xlsx"
first_file_list_global = []
second_file_list_global = []
third_file_list_global = []
files = [[first_file, first_file_list_global], [second_file, second_file_list_global],
         [third_file, third_file_list_global]]
# загружаем в память уже существующий файл на диске
# workbook1 = openpyxl.load_workbook(file_name)
#
# worksheet1 = workbook1.active
glob_list = []

for i in files:
    workbook = openpyxl.load_workbook(i[0])
    worksheet = workbook.active
    for y in range(1, 26):
        mass_local = []
        for x in range(1, 26):
            value = worksheet.cell(row=x, column=y).value
            if value is None:
                break
            mass_local.append(value)
        glob_list.append(mass_local)


    # filter(None, i[1])
    # i[1] = [x for x in i[1] if x]

    def iii():
        for ii in glob_list:
            # print("ii" + str(ii[0]))
            if len(ii) == 0:
                # print("ii")
                del glob_list[0]
                iii()

file_name = "task2.xlsx"

wb = Workbook()
# ws0 = wb.create_sheet("Лист1")
ind_element = 1
for element in glob_list:
    if len(element) == 0:
        continue
    else:
        ws = wb.create_sheet("Лист" + str(ind_element))
        row = 0
        col = 1
        while True:
            row += 1
            ws.cell(row=row, column=col).value = str(element[row-1])
            if row == len(element):
                break

        ind_element += 1
# Вишенка
if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])
# Вишенка
wb.save("task2.xlsx")
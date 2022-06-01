import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_interval

file_name = "exel.xlsx"

# загружаем в память уже существующий файл на диске
workbook1 = openpyxl.load_workbook(file_name)

worksheet1 = workbook1.active

mass_global = []


for y in range(1, 26):
    mass_local = []
    for x in range(1, 26):
        value = worksheet1.cell(row=x, column=y).value
        if value is None:
            break
        mass_local.append(value)
    mass_global.append(mass_local)
    print(mass_global)

workbook2 = Workbook()
worksheet2 = workbook2.active

row = 0
col = 1
for i in mass_global:
    row += 1
    for ii in i:
        print(ii)
        worksheet2.cell(row=row, column=col).value = str(ii)
        print(type(worksheet2[f'{col}{row}'][0]))
        col += 1
    col = 1

workbook2.save("sample_example.xlsx")
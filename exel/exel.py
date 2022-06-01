from openpyxl import Workbook
from openpyxl import load_workbook
import datetime

wb = load_workbook("exel.xlsx")

ws = wb.active

# ws['A1'] = 42

first = ws['A1:Y1']
# ws.append([10, 2, 3])
print(first)

# ws['A2'] = datetime.datetime.now()

# wb.save("sample.xlsx")
